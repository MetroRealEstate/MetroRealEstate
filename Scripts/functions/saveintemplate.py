import os
import openpyxl

def save_in_template(project_block):
    copied_excel_path = "COPIA PLANTILLA.xlsx"  # Change this path to the desired output path for the copied file

    # Load the existing "COPIA PLANTILLA.xlsx" if it exists, otherwise create a new workbook
    if os.path.exists(copied_excel_path):
        copied_workbook = openpyxl.load_workbook(copied_excel_path)
    else:
        copied_workbook = openpyxl.Workbook()

    # Select the first sheet in the copied workbook
    copied_sheet = copied_workbook.active

    # Define the starting row to write the data (row 5)
    start_row = 5

    # Find the first empty row in the copied sheet after the formatted section
    empty_row = start_row
    while copied_sheet.cell(row=empty_row, column=1).value is not None:
        empty_row += 1

    # Write headers in copied Excel if the sheet is empty
    if empty_row == start_row:
        copied_headers = list(project_block[0].keys())
        for col_num, header in enumerate(copied_headers, start=1):
            copied_sheet.cell(row=start_row, column=col_num, value=header)

    # Write data rows in copied Excel starting from the first empty row after the formatted section
    for item in project_block:
        row_data = list(item.values())
        for col_num, cell_value in enumerate(row_data, start=1):
            copied_sheet.cell(row=empty_row, column=col_num, value=cell_value)
        empty_row += 1

    # Save the copied Excel file
    copied_workbook.save(copied_excel_path)
